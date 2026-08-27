from openpyxl.formatting.rule import CellIsRule, FormulaRule
# from openpyxl.styles import PatternFill
from openpyxl.styles import Font

from exporter.excel_conditional_formats import ColumnConditionalFormat
from exporter.excel_formulas import FormulaColumn, SummaryFormula, OverallFormula

from config.data_config import NAME_COL, TARGET_MONTH


SHEET_SPLIT_COL = NAME_COL
OUTPUT_FILE_NAME = f"output_{TARGET_MONTH.replace('-', '')}.xlsx"


def payroll_lookup_summary(label, start_row, default=0):
    def build_formula(ctx):
        return (
            f"=INDEX(PayrollTable[{label}], "
            f"MATCH(INDEX({ctx.column_range('이름')},1), "
            f"PayrollTable[이름], 0))"
        )

    return SummaryFormula(
        label=label,
        formula=build_formula,
        number_format="#,##0",
        start_row=start_row,
    )

personal_formula_columns = [
    FormulaColumn(
        header="총근무시간",
        formula=lambda ctx: (
            f"=INT(ROUND(({ctx.cell('실퇴근시간')}-{ctx.cell('실출근시간')})*24, 6))"
        ),
        number_format="0"
    ),
    FormulaColumn(
        header="실근무시간",
        formula=lambda ctx: (
            f"=MAX({ctx.cell('총근무시간')} - 1, 0)"
        ),
        number_format="0"
    ),
    FormulaColumn(
        header="소정근무시간",
        formula=lambda ctx: (
            f"=IF({ctx.cell('휴일')} = 1, 0, MIN(MAX({ctx.cell('실근무시간')}, 0), 8))"
        ),
        number_format="0"
    ),
    FormulaColumn(
        header="평일연장근무시간",
        formula=lambda ctx: (
            f"=IF({ctx.cell('휴일')} = 1, 0, MAX({ctx.cell('실근무시간')} - {ctx.cell('소정근무시간')}, 0))"
        ),
        number_format="0"
    ),
    FormulaColumn(
        header="휴일근무시간",
        formula=lambda ctx: (
            f"=IF({ctx.cell('휴일')} = 1, MIN(MAX({ctx.cell('실근무시간')}, 0), 8), 0)"
        ),
        number_format="0"
    ),
    FormulaColumn(
        header="휴일연장근무시간",
        formula=lambda ctx: (
            f"=IF({ctx.cell('휴일')} = 1, MAX({ctx.cell('실근무시간')} - {ctx.cell('휴일근무시간')}, 0), 0)"
        ),
        number_format="0"
    ),
    FormulaColumn(
        header="야간근무시간",
        formula=lambda ctx: (
            f"=MAX(0, MIN({ctx.cell('실퇴근시간')}, INT({ctx.cell('실출근시간')}) - (MOD({ctx.cell('실출근시간')}, 1) < 6/24) + 30/24) - MAX({ctx.cell('실출근시간')}, INT({ctx.cell('실출근시간')}) - (MOD({ctx.cell('실출근시간')}, 1) < 6/24) + 22/24)) * 24"
        ),
        number_format="0"
    ),
    FormulaColumn(
        header="주휴인정시간",
        formula=lambda ctx: (
            f'=IFERROR(IF(AND('
            f'{ctx.cell("요일")}="일",'
            f'COUNTIF(OFFSET({ctx.cell("소정근무시간")},-6,0,5,1),">0")=5'
            f'),8,0),0)'
        ),
        number_format="0",
    ),
    FormulaColumn(
        header="공휴일인정시간",
        formula=lambda ctx: f'=IF({ctx.cell("공휴일")} = 1, 8, 0)',
        number_format="0",
    )
]

personal_summary_formulas = [
    payroll_lookup_summary("통상시급", start_row=2, default=10320),
    SummaryFormula(
        label="소정근무시간합계",
        formula=lambda ctx: (
            f"=SUM({ctx.column_range('소정근무시간')})"
        ),
        number_format="0",
        start_row=4
    ),
    SummaryFormula(
        label="주휴인정시간",
        formula=lambda ctx: (
            f"=SUM({ctx.column_range('주휴인정시간')})"
        ),
        number_format="0",
        start_row=5
    ),
    SummaryFormula(
        label="평일연장근무시간합계",
        formula=lambda ctx: (
            f"=SUM({ctx.column_range('평일연장근무시간')})"
        ),
        number_format="0",
        start_row=6
    ),
    SummaryFormula(
        label="휴일근무시간합계",
        formula=lambda ctx: (
            f"=SUM({ctx.column_range('휴일근무시간')})"
        ),
        number_format="0",
        start_row=7
    ),
    SummaryFormula(
        label="휴일연장근무시간합계",
        formula=lambda ctx: (
            f"=SUM({ctx.column_range('휴일연장근무시간')})"
        ),
        number_format="0",
        start_row=8
    ),
    SummaryFormula(
        label="야간근무시간합계",
        formula=lambda ctx: (
            f"=SUM({ctx.column_range('야간근무시간')})"
        ),
        number_format="0",
        start_row=9
    ),
    SummaryFormula(
        label="기본급",
        formula=lambda ctx: (
            f"=( {ctx.summary_cell('소정근무시간합계')} + {ctx.summary_cell('주휴인정시간')} ) * {ctx.summary_cell('통상시급')}"
        ),
        number_format="#,##0",
        start_row=12,
    ),
    SummaryFormula(
        label="평일연장수당",
        formula=lambda ctx: (
            f"={ctx.summary_cell('평일연장근무시간합계')} * {ctx.summary_cell('통상시급')} * 1.5"
        ),
        number_format="#,##0",
        start_row=13,
    ),
    SummaryFormula(
        label="휴일근무수당",
        formula=lambda ctx: (
            f"={ctx.summary_cell('휴일근무시간합계')} * {ctx.summary_cell('통상시급')} * 1.5"
        ),
        number_format="#,##0",
        start_row=14,
    ),
    SummaryFormula(
        label="휴일연장수당",
        formula=lambda ctx: (
            f"={ctx.summary_cell('휴일연장근무시간합계')} * {ctx.summary_cell('통상시급')} * 2.0"
        ),
        number_format="#,##0",
        start_row=15,
    ),
    SummaryFormula(
        label="야간수당",
        formula=lambda ctx: (
            f"={ctx.summary_cell('야간근무시간합계')} * {ctx.summary_cell('통상시급')} * 0.5"
        ),
        number_format="#,##0",
        start_row=16,
    ),
    SummaryFormula(
        label="기타수당",
        formula=lambda ctx: (
            f"=0"
        ),
        number_format="#,##0",
        start_row=17,
    ),
    SummaryFormula(
        label="지급합계",
        formula=lambda ctx: (
            f"=SUM({ctx.summary_cell('기본급')},{ctx.summary_cell('평일연장수당')},{ctx.summary_cell('휴일근무수당')},{ctx.summary_cell('휴일연장수당')},{ctx.summary_cell('야간수당')},{ctx.summary_cell('기타수당')})"
        ),
        number_format="#,##0",
        start_row=18,
    ),
    *[
        payroll_lookup_summary(label, start_row)
        for label, start_row in [
            ("고용보험", 20),
            ("고용보험정산", 21),
            ("국민연금", 22),
            ("건강보험", 23),
            ("건강보험정산", 24),
            ("장기요양", 25),
            ("장기요양정산", 26),
            ("환급금이자", 27),
            ("관리비", 28),
            ("식대비", 29),
            ("소득세", 30),
            ("지방소득세", 31),
            ("지방소득세정산", 32),
            ("기타공제", 33),
        ]
    ],
    SummaryFormula(
        label="공제합계",
        formula=lambda ctx: (
            f"=SUM({ctx.summary_cell('고용보험')},{ctx.summary_cell('고용보험정산')},{ctx.summary_cell('국민연금')},{ctx.summary_cell('건강보험')},{ctx.summary_cell('건강보험정산')},{ctx.summary_cell('장기요양')},{ctx.summary_cell('장기요양정산')},{ctx.summary_cell('환급금이자')},{ctx.summary_cell('관리비')},{ctx.summary_cell('식대비')},{ctx.summary_cell('소득세')},{ctx.summary_cell('지방소득세')},{ctx.summary_cell('지방소득세정산')},{ctx.summary_cell('기타공제')})"
        ),
        number_format="#,##0",
        start_row=34,
    ),
]

overall_formula_columns = [
    OverallFormula(
        header="소정근무",
        formula=lambda ctx: ctx.vlookup_summary(
            "사용자ID",
            "소정근무시간합계",
            default=0,
        ),
        number_format="0",
    ),
    OverallFormula(
        header="주휴",
        formula=lambda ctx: ctx.vlookup_summary(
            "사용자ID",
            "주휴인정시간",
            default=0,
        ),
        number_format="0",
    ),

    OverallFormula(
        header="평일연장",
        formula=lambda ctx: ctx.vlookup_summary(
            "사용자ID",
            "평일연장근무시간합계",
            default=0,
        ),
        number_format="0",
    ),
    OverallFormula(
        header="휴일근무",
        formula=lambda ctx: ctx.vlookup_summary(
            "사용자ID",
            "휴일근무시간합계",
            default=0,
        ),
        number_format="0",
    ),
    OverallFormula(
        header="휴일연장근무",
        formula=lambda ctx: ctx.vlookup_summary(
            "사용자ID",
            "휴일연장근무시간합계",
            default=0,
        ),
        number_format="0",
    ),
    OverallFormula(
        header="야간근무",
        formula=lambda ctx: ctx.vlookup_summary(
            "사용자ID",
            "야간근무시간합계",
            default=0,
        ),
        number_format="0",
    ),

    OverallFormula(
        header="기본급",
        formula=lambda ctx: ctx.vlookup_summary(
            "사용자ID",
            "기본급",
            default=0,
        ),
        number_format="0",
    ),
    OverallFormula(
        header="평일연장수당",
        formula=lambda ctx: ctx.vlookup_summary(
            "사용자ID",
            "평일연장수당",
            default=0,
        ),
        number_format="0",
    ),
    OverallFormula(
        header="휴일근무수당",
        formula=lambda ctx: ctx.vlookup_summary(
            "사용자ID",
            "휴일근무수당",
            default=0,
        ),
        number_format="0",
    ),
    OverallFormula(
        header="휴일연장수당",
        formula=lambda ctx: ctx.vlookup_summary(
            "사용자ID",
            "휴일연장수당",
            default=0,
        ),
        number_format="0",
    ),

    OverallFormula(
        header="야간수당",
        formula=lambda ctx: ctx.vlookup_summary(
            "사용자ID",
            "야간수당",
            default=0,
        ),
        number_format="0",
    ),

    OverallFormula(
        header="기타수당",
        formula=lambda ctx: ctx.vlookup_summary(
            "사용자ID",
            "기타수당",
            default=0,
        ),
        number_format="0",
    ),

    OverallFormula(
        header="지급합계",
        formula=lambda ctx: ctx.vlookup_summary(
            "사용자ID",
            "지급합계",
            default=0,
        ),
        number_format="0",
    ),
]


# 조건부서식
red_font = Font(
    color="FFFF0000",  # 순수 빨간색 (또는 엑셀 기본 진한 빨강: "FF9C0006")
    bold=True          # (선택) 굵게 표시하고 싶을 경우
)
personal_conditional_formats = [
    ColumnConditionalFormat(
        column="근무일자",
        rule=FormulaRule(
            formula=[
                'COUNTIF('
                'INDEX($2:$1000, 0, MATCH("근무일자", $1:$1, 0)),'
                'INDEX(2:2, 1, MATCH("근무일자", $1:$1, 0))'
                ') > 1'
            ],
            font=red_font
        )
    ),
    ColumnConditionalFormat(
        column="총근무시간",
        rule=CellIsRule(
            operator="greaterThan",
            formula=["22"],
            font=red_font
        ),
    ),
    ColumnConditionalFormat(
        column="총근무시간",
        rule=CellIsRule(
            operator="lessThan",
            formula=["-1"],
            font=red_font
        ),
    ),
    ColumnConditionalFormat(
        column="실근무시간",
        rule=FormulaRule(
            formula=[
                'AND('
                'INDEX(2:2,1,MATCH("휴일",$1:$1,0))=0,'
                'INDEX(2:2,1,MATCH("교대일",$1:$1,0))=0,'
                'INDEX(2:2,1,MATCH("야간근무시간",$1:$1,0))>=8,'
                'INDEX(2:2,1,MATCH("실근무시간",$1:$1,0))>=13'
                ')'
            ],
            font=red_font,
        ),
    ),
    ColumnConditionalFormat(
        column="실근무시간",
        rule=FormulaRule(
            formula=[
                'AND('
                'INDEX(2:2,1,MATCH("휴일",$1:$1,0))=0,'
                'INDEX(2:2,1,MATCH("교대일",$1:$1,0))=0,'
                'INDEX(2:2,1,MATCH("야간근무시간",$1:$1,0))=0,'
                'INDEX(2:2,1,MATCH("실근무시간",$1:$1,0))>=12'
                ')'
            ],
            font=red_font,
        ),
    ),
    ColumnConditionalFormat(
        column="주휴인정시간",
        rule=FormulaRule(
            formula=[
                'AND('
                'INDEX(2:2,1,MATCH("주휴인정시간",$1:$1,0))=0,'
                'INDEX(2:2,1,MATCH("요일",$1:$1,0))="일",'
                'COUNTIFS(OFFSET(INDEX(1:1,1,MATCH("요일",$1:$1,0)),1,0,ROW()-1,1),"일")=1'
                ')'
            ],
            font=red_font,
        ),
    ),
    ColumnConditionalFormat(
        column="근무일자",
        rule=FormulaRule(
            formula=[
                'INDEX(2:2,1,MATCH("공휴일",$1:$1,0))=1'
            ],
            font=red_font,
        ),
    )
]