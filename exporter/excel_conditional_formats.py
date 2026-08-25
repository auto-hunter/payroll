"""열 이름을 기준으로 Excel 조건부서식을 적용하는 모듈."""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from typing import Iterable

from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class ColumnConditionalFormat:
    """데이터 열 하나에 적용할 조건부서식 규칙."""

    column: str
    rule: Rule


def apply_column_conditional_formats(
    worksheet: Worksheet,
    conditional_formats: Iterable[ColumnConditionalFormat],
    *,
    header_row: int = 1,
) -> None:
    """열 이름으로 데이터 범위를 찾아 조건부서식을 적용한다."""
    rules = tuple(conditional_formats)
    if not rules or worksheet.max_row <= header_row:
        return

    column_indexes = {
        str(worksheet.cell(header_row, column).value): column
        for column in range(1, worksheet.max_column + 1)
        if worksheet.cell(header_row, column).value is not None
    }

    for conditional_format in rules:
        column_name = str(conditional_format.column).strip()
        if not column_name:
            raise ValueError("조건부서식을 적용할 열 이름은 비어 있을 수 없습니다.")
        if column_name not in column_indexes:
            raise KeyError(f"{column_name} 컬럼을 조건부서식에서 찾을 수 없습니다.")

        column_letter = get_column_letter(column_indexes[column_name])
        cell_range = (
            f"{column_letter}{header_row + 1}:"
            f"{column_letter}{worksheet.max_row}"
        )
        # 동일한 설정을 여러 워크시트에 적용하므로 Rule 객체를 복사한다.
        worksheet.conditional_formatting.add(
            cell_range,
            copy(conditional_format.rule),
        )
