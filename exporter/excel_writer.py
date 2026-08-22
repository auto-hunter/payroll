"""openpyxl을 이용한 엑셀 파일 출력 모듈.

DataFrame을 이름별 시트로 나누어 기록한다.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.cell import absolute_coordinate, get_column_letter
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.table import Table

from config.config import COMPANY_COL, SHEET_SPLIT_COL
from exporter.excel_formulas import (
    FormulaColumn,
    OverallFormula,
    OverallFormulaContext,
    SummaryFormula,
    apply_formula_columns,
    apply_summary_formulas,
)


def _make_sheet_title(value: object, used_titles: set[str]) -> str:
    """그룹 값을 엑셀 규칙에 맞는 고유한 시트명으로 변환한다.

    엑셀 시트명에는 일부 문자를 사용할 수 없고 최대 길이는 31자이다.
    또한 대소문자만 다른 이름도 같은 시트명으로 취급되므로, 이미 사용한
    이름과 겹치면 ``_2``, ``_3``과 같은 번호를 붙인다.

    Args:
        value: 시트명의 원본이 되는 그룹 값(기본적으로 직원 이름).
        used_titles: 현재 워크북에서 이미 사용한 시트명의 소문자 집합.

    Returns:
        엑셀에서 사용할 수 있으며 워크북 안에서 중복되지 않는 시트명.
    """
    # 이름이 비어 있거나 NaN인 그룹도 시트를 만들 수 있게 기본명을 준다.
    if pd.isna(value):
        base_title = "이름없음"
    else:
        base_title = str(value).strip() or "이름없음"

    # 엑셀이 금지하는 문자와 제어 문자를 '_'로 바꾸고 31자로 자른다.
    base_title = re.sub(r"[\\/*?:\[\]\x00-\x1f]", "_", base_title)[:31]
    title = base_title
    sequence = 2

    # casefold()를 사용해 영문 대소문자가 다른 제목도 중복으로 판단한다.
    while title.casefold() in used_titles:
        suffix = f"_{sequence}"
        # 번호까지 포함한 최종 제목도 31자를 넘지 않도록 원래 이름을 줄인다.
        title = f"{base_title[:31 - len(suffix)]}{suffix}"
        sequence += 1

    used_titles.add(title.casefold())
    return title


def _to_excel_value(value: object) -> object:
    """pandas/numpy 값을 openpyxl이 셀에 기록할 수 있는 값으로 변환한다.

    결측값은 빈 셀을 뜻하는 ``None``으로, pandas 날짜는 파이썬 datetime으로,
    numpy 스칼라는 파이썬 기본 스칼라로 변환한다. 그 밖의 값은 그대로 둔다.

    Args:
        value: DataFrame의 셀 값.

    Returns:
        openpyxl의 ``Worksheet.append``에 전달할 수 있는 값.
    """
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    # numpy.int64 같은 numpy 스칼라는 item()으로 int 등의 기본형이 된다.
    if hasattr(value, "item"):
        return value.item()
    return value


def _write_dataframe(worksheet, df: pd.DataFrame) -> None:
    """DataFrame의 헤더와 행을 워크시트에 기록한다."""
    worksheet.append([str(column) for column in df.columns])
    for row in df.itertuples(index=False, name=None):
        worksheet.append([_to_excel_value(value) for value in row])


def _lookup_key(value: object) -> str:
    """시트 분리 값과 급여대장 행을 비교하기 위한 키를 만든다."""
    return "<이름없음>" if pd.isna(value) else str(value)


def _quote_sheet_title(title: str) -> str:
    """엑셀 수식에서 사용할 수 있도록 시트명의 작은따옴표를 이스케이프한다."""
    return title.replace("'", "''")


def write_dataframe_by_name(
    df: pd.DataFrame,
    output_path: str | Path,
    sheet_split_col: str = SHEET_SPLIT_COL,
    personal_formula_columns: Iterable[FormulaColumn] = (),
    personal_summary_formulas: Iterable[SummaryFormula] = (),
    df_overall: pd.DataFrame | None = None,
    overall_sheet_name: str = "급여대장",
    overall_summary_labels: Iterable[str] = (),
    overall_formula_columns: Iterable[OverallFormula] = (),
    user_id_col: str = "사용자ID",
    company_col: str = COMPANY_COL,
    df_deductions: pd.DataFrame | None = None,
    deduction_join_col: str = "사용자ID",
) -> Path:
    """DataFrame을 이름별 시트로 나누어 새 엑셀 파일에 저장한다.

    Args:
        df: 저장할 전체 데이터.
        output_path: 생성할 xlsx 파일의 경로.
        sheet_split_col: 데이터를 나누는 기준 컬럼.
        personal_formula_columns: 개인별 시트에 추가할 수식 열 규칙.
        personal_summary_formulas: 개인별 시트의 데이터 우측에 추가할 요약 수식 규칙.
        df_overall: 급여대장의 기본 데이터. 생략하면 분리 기준 컬럼만 생성한다.
        overall_sheet_name: 첫 번째 시트로 만들 급여대장의 이름.
        overall_summary_labels: 개인별 요약 셀을 참조해 급여대장에 추가할 항목명.
        overall_formula_columns: 사용자ID로 개인 요약 값을 조회할 급여대장 수식 열 규칙.
        user_id_col: 개인 시트와 급여대장을 연결하는 사용자ID 컬럼.
        company_col: 개인 시트에서 급여대장으로 가져올 등록사업장 컬럼.
        df_deductions: 급여대장 수식 열 뒤에 추가할 공제 정보.
        deduction_join_col: 급여대장과 공제 정보를 연결할 컬럼.

    Returns:
        저장을 마친 파일의 ``Path`` 객체.

    Raises:
        ValueError: DataFrame이 없거나 비어 있을 때.
        KeyError: 그룹 기준 컬럼이 DataFrame에 없을 때.
    """
    # 빈 데이터로는 시트를 만들 수 없으므로 파일 생성 전에 명확히 실패시킨다.
    if df is None or df.empty:
        raise ValueError("저장할 데이터가 없습니다.")
    if sheet_split_col not in df.columns:
        raise KeyError(f"{sheet_split_col} 컬럼이 없습니다.")

    workbook = Workbook()

    # openpyxl은 수식을 기록할 수는 있지만 계산 엔진은 제공하지 않는다.
    # 따라서 Excel에서 파일을 열 때 모든 수식을 자동으로 다시 계산하도록
    # 통합문서 계산 속성을 명시한다. calcId=0은 현재 Excel 버전의 계산
    # 엔진으로 수식 캐시를 새로 만들도록 유도한다.
    workbook.calculation = CalcProperties(
        calcMode="auto",
        calcId=0,
        fullCalcOnLoad=True,
        forceFullCalc=True,
        calcOnSave=True,
    )

    # Workbook이 자동 생성하는 빈 기본 시트를 제거한다.
    workbook.remove(workbook.active)
    used_titles: set[str] = set()
    # generator로 전달된 규칙도 모든 시트에서 재사용할 수 있게 고정한다.
    personal_formula_rules = tuple(personal_formula_columns)
    personal_summary_rules = tuple(personal_summary_formulas)
    overall_formula_rules = tuple(overall_formula_columns)
    needs_user_id = bool(overall_formula_rules) or df_deductions is not None
    overall_labels = tuple(str(label).strip() for label in overall_summary_labels)
    if any(not label for label in overall_labels):
        raise ValueError("급여대장 요약 항목명은 비어 있을 수 없습니다.")
    if len(set(overall_labels)) != len(overall_labels):
        raise ValueError("급여대장 요약 항목명이 중복되었습니다.")

    summary_rule_by_label = {
        str(rule.label).strip(): rule for rule in personal_summary_rules
    }
    unknown_labels = [
        label for label in overall_labels if label not in summary_rule_by_label
    ]
    if unknown_labels:
        raise KeyError(
            f"개인별 요약 수식에 없는 항목입니다: {', '.join(unknown_labels)}"
        )

    # 급여대장은 개인 시트의 실제 셀 주소를 참조하므로 개인 시트를 먼저
    # 만든 뒤 마지막에 생성하고 첫 번째 위치로 이동한다.
    overall_title = _make_sheet_title(overall_sheet_name, used_titles)
    personal_summaries: dict[str, tuple[str, dict[str, str]]] = {}
    personal_summaries_by_user_id: dict[
        str, tuple[str, str, dict[str, str]]
    ] = {}
    split_values: list[object] = []
    split_user_ids: list[object] = []
    split_companies: list[object] = []

    # sort=False는 원본에 등장한 이름 순서를 유지하고, dropna=False는
    # 이름이 비어 있는 행도 누락하지 않고 '이름없음' 시트에 포함한다.
    for name, df_name in df.groupby(sheet_split_col, dropna=False, sort=False):
        worksheet = workbook.create_sheet(_make_sheet_title(name, used_titles))
        _write_dataframe(worksheet, df_name)

        # 데이터 오른쪽에 호출자가 정의한 수식 열을 동일하게 추가한다.
        apply_formula_columns(worksheet, personal_formula_rules)
        # 빈 열 하나를 사이에 두고 제목/값 형태의 요약 영역을 추가한다.
        summary_cells = apply_summary_formulas(worksheet, personal_summary_rules)
        key = _lookup_key(name)
        personal_summaries[key] = (worksheet.title, summary_cells)
        split_values.append(name)

        if df_overall is None:
            if company_col not in df_name.columns:
                raise KeyError(f"개인별 데이터에 {company_col} 컬럼이 없습니다.")
            raw_companies = df_name[company_col].dropna().unique()
            if len(raw_companies) > 1:
                raise ValueError(
                    f"{worksheet.title} 시트에는 하나의 {company_col} 값만 있어야 합니다."
                )
            split_companies.append(
                raw_companies[0] if len(raw_companies) == 1 else None
            )

        if needs_user_id:
            if user_id_col not in df_name.columns:
                raise KeyError(f"개인별 데이터에 {user_id_col} 컬럼이 없습니다.")
            raw_user_ids = df_name[user_id_col].dropna().unique()
            user_ids = [_lookup_key(value) for value in raw_user_ids]
            if len(user_ids) != 1:
                raise ValueError(
                    f"{worksheet.title} 시트에는 하나의 {user_id_col} 값만 있어야 합니다."
                )
            user_id_key = str(user_ids[0])
            if user_id_key in personal_summaries_by_user_id:
                raise ValueError(f"{user_id_col} 값이 중복되었습니다: {user_id_key}")
            user_id_column = df_name.columns.get_loc(user_id_col) + 1
            user_id_cell = absolute_coordinate(
                f"{get_column_letter(user_id_column)}2"
            )
            personal_summaries_by_user_id[user_id_key] = (
                worksheet.title,
                user_id_cell,
                summary_cells,
            )
            split_user_ids.append(raw_user_ids[0])

    if df_overall is None:
        overall_columns = {sheet_split_col: split_values}
        if needs_user_id:
            overall_columns[user_id_col] = split_user_ids
        overall_columns[company_col] = split_companies
        overall_data = pd.DataFrame(overall_columns)
    else:
        overall_data = df_overall.copy()
        if sheet_split_col not in overall_data.columns:
            raise KeyError(f"급여대장에 {sheet_split_col} 컬럼이 없습니다.")

    if overall_data[sheet_split_col].map(_lookup_key).duplicated().any():
        raise ValueError(f"급여대장의 {sheet_split_col} 값은 중복될 수 없습니다.")

    overall_keys = set(overall_data[sheet_split_col].map(_lookup_key))
    missing_keys = [key for key in personal_summaries if key not in overall_keys]
    if missing_keys:
        raise ValueError(
            f"급여대장에 없는 {sheet_split_col} 값입니다: {', '.join(missing_keys)}"
        )

    overlapping_labels = [
        label for label in overall_labels if label in overall_data.columns
    ]
    if overlapping_labels:
        raise ValueError(
            "급여대장 기본 데이터와 개인 요약 항목이 중복됩니다: "
            + ", ".join(overlapping_labels)
        )

    overall_worksheet = workbook.create_sheet(overall_title, 0)
    _write_dataframe(overall_worksheet, overall_data)

    for label in overall_labels:
        column = overall_worksheet.max_column + 1
        overall_worksheet.cell(1, column, label)
        rule = summary_rule_by_label[label]
        for row in range(2, overall_worksheet.max_row + 1):
            value = overall_worksheet.cell(
                row, overall_data.columns.get_loc(sheet_split_col) + 1
            ).value
            personal = personal_summaries.get(_lookup_key(value))
            if personal is None:
                continue
            sheet_title, summary_cells = personal
            if label not in summary_cells:
                raise KeyError(f"{sheet_title} 시트에 {label} 요약 셀이 없습니다.")
            source_cell = absolute_coordinate(summary_cells[label])
            formula = f"='{_quote_sheet_title(sheet_title)}'!{source_cell}"
            cell = overall_worksheet.cell(row, column, formula)
            if rule.number_format is not None:
                cell.number_format = rule.number_format

    if overall_formula_rules:
        if user_id_col not in overall_data.columns:
            raise KeyError(f"급여대장에 {user_id_col} 컬럼이 없습니다.")

        overall_columns = {
            str(overall_worksheet.cell(1, index).value): index
            for index in range(1, overall_worksheet.max_column + 1)
        }
        user_id_column = overall_columns[user_id_col]

        for rule in overall_formula_rules:
            header = str(rule.header).strip()
            if not header:
                raise ValueError("급여대장 수식 열 이름은 비어 있을 수 없습니다.")
            if header in overall_columns:
                raise ValueError(f"급여대장에 {header} 컬럼이 이미 존재합니다.")

            target_column = overall_worksheet.max_column + 1
            overall_worksheet.cell(1, target_column, header)
            overall_columns[header] = target_column

            for row in range(2, overall_worksheet.max_row + 1):
                user_id = overall_worksheet.cell(row, user_id_column).value
                personal = personal_summaries_by_user_id.get(_lookup_key(user_id))
                if personal is None:
                    continue
                sheet_title, personal_user_id_cell, summary_cells = personal
                context = OverallFormulaContext(
                    row=row,
                    columns=overall_columns,
                    personal_sheet_title=sheet_title,
                    personal_user_id_cell=personal_user_id_cell,
                    personal_summary_cells=summary_cells,
                )
                formula = rule.formula(context)
                if not isinstance(formula, str) or not formula.startswith("="):
                    raise ValueError(
                        f"{header} 컬럼의 {row}행 수식은 '='으로 시작해야 합니다."
                    )
                cell = overall_worksheet.cell(row, target_column, formula)
                if rule.number_format is not None:
                    cell.number_format = rule.number_format

    # 개인 요약 수식 열이 모두 만들어진 뒤 공제 정보를 가장 오른쪽에 붙인다.
    if df_deductions is not None:
        if deduction_join_col not in overall_data.columns:
            raise KeyError(f"급여대장에 {deduction_join_col} 컬럼이 없습니다.")
        if deduction_join_col not in df_deductions.columns:
            raise KeyError(f"공제정보에 {deduction_join_col} 컬럼이 없습니다.")
        if df_deductions.columns.duplicated().any():
            raise ValueError("공제정보의 컬럼명은 중복될 수 없습니다.")
        if df_deductions[deduction_join_col].isna().any():
            raise ValueError(f"공제정보의 {deduction_join_col} 값은 비어 있을 수 없습니다.")

        deduction_keys = df_deductions[deduction_join_col].map(_lookup_key)
        if deduction_keys.duplicated().any():
            raise ValueError(
                f"공제정보의 {deduction_join_col} 값은 중복될 수 없습니다."
            )

        deduction_columns = [
            column
            for column in df_deductions.columns
            if column != deduction_join_col
        ]
        existing_headers = {
            str(overall_worksheet.cell(1, column).value)
            for column in range(1, overall_worksheet.max_column + 1)
        }
        duplicated_headers = [
            str(column)
            for column in deduction_columns
            if str(column) in existing_headers
        ]
        if duplicated_headers:
            raise ValueError(
                "급여대장과 공제정보의 컬럼명이 중복됩니다: "
                + ", ".join(duplicated_headers)
            )

        deduction_lookup = {
            _lookup_key(row[deduction_join_col]): row
            for _, row in df_deductions.iterrows()
        }
        overall_join_column = overall_data.columns.get_loc(deduction_join_col) + 1
        first_deduction_column = overall_worksheet.max_column + 1

        for offset, deduction_column in enumerate(deduction_columns):
            target_column = first_deduction_column + offset
            overall_worksheet.cell(1, target_column, str(deduction_column))

            for row_number in range(2, overall_worksheet.max_row + 1):
                join_value = overall_worksheet.cell(
                    row_number, overall_join_column
                ).value
                deduction_row = deduction_lookup.get(_lookup_key(join_value))
                if deduction_row is None:
                    continue
                overall_worksheet.cell(
                    row_number,
                    target_column,
                    _to_excel_value(deduction_row[deduction_column]),
                )

    # 모든 열이 만들어진 급여대장 데이터 영역을 Excel 테이블로 지정한다.
    table_ref = (
        f"A1:{get_column_letter(overall_worksheet.max_column)}"
        f"{overall_worksheet.max_row}"
    )
    overall_worksheet.add_table(Table(displayName="PayrollTable", ref=table_ref))

    destination = Path(output_path)
    # 중간 폴더가 없어도 저장할 수 있도록 상위 경로를 먼저 생성한다.
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination
