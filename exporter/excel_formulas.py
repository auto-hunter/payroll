"""워크시트에 행별 수식 열과 요약 수식 영역을 추가하는 모듈.

각 열의 헤더와 수식 생성 규칙을 :class:`FormulaColumn`로 정의하고,
요약 항목은 :class:`SummaryFormula`로 정의한다.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Callable, Iterable, Mapping

from openpyxl.utils import absolute_coordinate, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class FormulaContext:
    """한 데이터 행의 수식을 만드는 데 필요한 셀 위치 정보."""

    row: int
    columns: Mapping[str, int]

    def cell(self, column_name: str) -> str:
        """컬럼명에 해당하는 현재 행의 셀 주소를 반환한다.

        예를 들어 ``row=2``이고 ``"출근시간"``이 C열이면 ``"C2"``를
        반환한다. 존재하지 않는 컬럼명에는 ``KeyError``가 발생한다.
        """
        try:
            column_index = self.columns[column_name]
        except KeyError as error:
            raise KeyError(f"{column_name} 컬럼을 수식에서 찾을 수 없습니다.") from error
        return f"{get_column_letter(column_index)}{self.row}"


@dataclass(frozen=True)
class SummaryContext:
    """워크시트의 데이터 범위를 참조하는 요약 수식 컨텍스트."""

    first_data_row: int
    last_data_row: int
    columns: Mapping[str, int]
    summary_cells: Mapping[str, str]

    def column_range(self, column_name: str) -> str:
        """컬럼명에 해당하는 전체 데이터 범위를 반환한다.

        예를 들어 데이터가 2~10행이고 ``"실근무시간"``이 H열이면
        ``"H2:H10"``을 반환한다.
        """
        try:
            column_index = self.columns[column_name]
        except KeyError as error:
            raise KeyError(f"{column_name} 컬럼을 요약 수식에서 찾을 수 없습니다.") from error

        column_letter = get_column_letter(column_index)
        return f"{column_letter}{self.first_data_row}:{column_letter}{self.last_data_row}"

    def summary_cell(self, label: str) -> str:
        """요약 제목에 해당하는 값 셀 주소를 반환한다.

        예를 들어 ``"소정근무시간합계"``의 값이 Q4에 기록된다면
        ``"Q4"``를 반환한다.
        """
        try:
            return self.summary_cells[label]
        except KeyError as error:
            raise KeyError(f"{label} 요약 항목을 찾을 수 없습니다.") from error


@dataclass(frozen=True)
class OverallFormulaContext:
    """급여대장 행에서 개인 시트의 요약 값을 조회하는 수식 컨텍스트."""

    row: int
    columns: Mapping[str, int]
    personal_sheet_title: str
    personal_user_id_cell: str
    personal_summary_cells: Mapping[str, str]

    def cell(self, column_name: str) -> str:
        """급여대장 현재 행에서 지정한 컬럼의 셀 주소를 반환한다."""
        try:
            column_index = self.columns[column_name]
        except KeyError as error:
            raise KeyError(
                f"{column_name} 컬럼을 급여대장 수식에서 찾을 수 없습니다."
            ) from error
        return f"{get_column_letter(column_index)}{self.row}"

    def vlookup_summary(
        self,
        user_id_column: str,
        summary_label: str,
        *,
        default: str | int | float = 0,
    ) -> str:
        """사용자ID로 개인 시트의 SummaryFormula 결과를 조회한다."""
        try:
            summary_cell = self.personal_summary_cells[summary_label]
        except KeyError as error:
            raise KeyError(
                f"{summary_label} 개인 요약 항목을 찾을 수 없습니다."
            ) from error

        sheet = self.personal_sheet_title.replace("'", "''")
        lookup_value = self.cell(user_id_column)
        user_id_cell = absolute_coordinate(self.personal_user_id_cell)
        summary_cell = absolute_coordinate(summary_cell)
        return (
            f"=IFERROR(VLOOKUP({lookup_value},"
            f"CHOOSE({{1,2}},'{sheet}'!{user_id_cell},"
            f"'{sheet}'!{summary_cell}),2,FALSE),{default})"
        )


FormulaBuilder = Callable[[FormulaContext], str]
SummaryFormulaBuilder = Callable[[SummaryContext], str]
OverallFormulaBuilder = Callable[[OverallFormulaContext], str]


@dataclass(frozen=True)
class FormulaColumn:
    """추가할 수식 열 하나의 규칙.

    Args:
        header: 새로 추가할 열 이름.
        formula: 각 데이터 행에서 엑셀 수식 문자열을 만드는 함수.
        number_format: 선택적인 엑셀 셀 표시 형식.
    """

    header: str
    formula: FormulaBuilder
    number_format: str | None = None


@dataclass(frozen=True)
class SummaryFormula:
    """워크시트 우측 요약 영역에 기록할 제목과 값 수식 규칙.

    Args:
        label: 요약 항목의 제목.
        formula: 데이터 범위를 이용해 값 셀의 수식을 만드는 함수.
        number_format: 값 셀에 적용할 선택적인 표시 형식.
        start_row: 이 요약 항목을 기록할 행. 생략하면 선언 순서에 따라
            헤더 다음 행부터 배치한다.
    """

    label: str
    formula: SummaryFormulaBuilder
    number_format: str | None = None
    start_row: int | None = None


@dataclass(frozen=True)
class OverallFormula:
    """급여대장에 추가할 개인 요약 조회 수식 열 규칙."""

    header: str
    formula: OverallFormulaBuilder
    number_format: str | None = None


def apply_formula_columns(
    worksheet: Worksheet,
    formula_columns: Iterable[FormulaColumn],
    *,
    header_row: int = 1,
) -> None:
    """워크시트 오른쪽에 지정된 수식 열들을 순서대로 추가한다.

    헤더는 ``header_row``에 기록하고, 그 다음 행부터 현재 워크시트의 마지막
    데이터 행까지 각 규칙이 만든 수식을 입력한다. 뒤쪽 규칙은 앞에서 추가한
    수식 열을 ``context.cell()``로 참조할 수도 있다.

    Raises:
        ValueError: 열 이름이 비어 있거나 기존/추가 열 이름과 중복될 때,
            또는 생성된 값이 ``=``으로 시작하는 엑셀 수식이 아닐 때.
    """
    rules = tuple(formula_columns)
    if not rules:
        return

    # 수식 열을 추가하기 전의 마지막 행을 보존한다. 헤더만 있는 시트에는
    # 수식 행을 만들지 않는다.
    last_data_row = worksheet.max_row
    column_indexes: dict[str, int] = {}

    for column_index in range(1, worksheet.max_column + 1):
        header = worksheet.cell(header_row, column_index).value
        if header is not None:
            column_indexes[str(header)] = column_index

    for rule in rules:
        header = str(rule.header).strip()
        if not header:
            raise ValueError("추가할 수식 열 이름은 비어 있을 수 없습니다.")
        if header in column_indexes:
            raise ValueError(f"{header} 컬럼이 이미 존재합니다.")

        target_column = worksheet.max_column + 1
        worksheet.cell(header_row, target_column, header)
        column_indexes[header] = target_column

        for row in range(header_row + 1, last_data_row + 1):
            context = FormulaContext(row=row, columns=column_indexes)
            formula = rule.formula(context)
            if not isinstance(formula, str) or not formula.startswith("="):
                raise ValueError(
                    f"{header} 컬럼의 {row}행 수식은 '='으로 시작해야 합니다."
                )

            cell = worksheet.cell(row, target_column, formula)
            if rule.number_format is not None:
                cell.number_format = rule.number_format


def apply_summary_formulas(
    worksheet: Worksheet,
    summary_formulas: Iterable[SummaryFormula],
    *,
    header_row: int = 1,
    spacer_columns: int = 1,
) -> dict[str, str]:
    """데이터 우측에 제목 열과 수식 값 열로 구성된 요약 영역을 추가한다.

    기존 데이터와 요약 제목 열 사이에는 기본적으로 빈 열 하나를 둔다.
    각 요약 수식은 헤더 다음 행부터 아래 방향으로 하나씩 기록된다.

    Args:
        worksheet: 요약 영역을 추가할 워크시트.
        summary_formulas: 순서대로 기록할 요약 수식 규칙.
        header_row: 원본 데이터의 헤더 행.
        spacer_columns: 기존 데이터와 요약 영역 사이의 빈 열 개수.

    Returns:
        요약 항목명을 해당 수식 값 셀 주소에 연결한 매핑.

    Raises:
        ValueError: 여백/시작 행이 잘못되었거나 제목 또는 수식이 유효하지
            않을 때.
    """
    rules = tuple(summary_formulas)
    if not rules:
        return {}
    if spacer_columns < 0:
        raise ValueError("spacer_columns는 0 이상이어야 합니다.")

    last_data_row = worksheet.max_row
    if last_data_row <= header_row:
        return {}

    column_indexes: dict[str, int] = {}
    for column_index in range(1, worksheet.max_column + 1):
        header = worksheet.cell(header_row, column_index).value
        if header is not None:
            column_indexes[str(header)] = column_index

    label_column = worksheet.max_column + spacer_columns + 1
    value_column = label_column + 1
    value_column_letter = get_column_letter(value_column)
    prepared_rules: list[tuple[SummaryFormula, str, int]] = []
    summary_cells: dict[str, str] = {}
    used_rows: set[int] = set()

    for offset, rule in enumerate(rules):
        label = str(rule.label).strip()
        if not label:
            raise ValueError("요약 항목 제목은 비어 있을 수 없습니다.")
        if label in summary_cells:
            raise ValueError(f"{label} 요약 항목이 중복되었습니다.")

        row = header_row + 1 + offset if rule.start_row is None else rule.start_row
        if row <= header_row:
            raise ValueError(f"{label}의 start_row는 header_row보다 커야 합니다.")
        if row in used_rows:
            raise ValueError(f"{row}행에 둘 이상의 요약 항목을 배치할 수 없습니다.")

        prepared_rules.append((rule, label, row))
        summary_cells[label] = f"{value_column_letter}{row}"
        used_rows.add(row)

    context = SummaryContext(
        first_data_row=header_row + 1,
        last_data_row=last_data_row,
        columns=column_indexes,
        summary_cells=summary_cells,
    )

    for rule, label, row in prepared_rules:
        formula = rule.formula(context)
        if not isinstance(formula, str) or not formula.startswith("="):
            raise ValueError(f"{label} 요약 수식은 '='으로 시작해야 합니다.")

        worksheet.cell(row, label_column, label)
        value_cell = worksheet.cell(row, value_column, formula)
        if rule.number_format is not None:
            value_cell.number_format = rule.number_format

    return summary_cells
